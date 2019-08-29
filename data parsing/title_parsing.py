
# coding: utf-8

# In[3]:


#https://blog.naver.com/godinus123/221421092663
#유튜브에서 제목 가져오기
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import time
import openpyxl


# In[5]:


driver = webdriver.Chrome('/Users/HYUNJIN BAE/Downloads/chromedriver')
driver.get('https://www.youtube.com/results?search_query=kbo+%EB%8B%A4%EC%8B%9C%EB%B3%B4%EA%B8%B0')

time.sleep(1)

body = driver.find_element_by_tag_name("body")
num_of_pagedowns =100 # 50
## 여러페이지를 유튜브에서 가져올때 100 페이지 가져오기 

while num_of_pagedowns:
    body.send_keys(Keys.PAGE_DOWN)
    time.sleep(2.3)
    num_of_pagedowns-=1
    try:
        driver.find_element_by_xpath().click()
    except:
        None

html = driver.page_source
soup = BeautifulSoup(html,'lxml')

    
arr_txt = list()

video_index=0
for sour in soup.find_all('a', href=True, title=True):
    tmp_str_0 = sour.get('aria-label')
    if(tmp_str_0 != None):
        T1 = tmp_str_0.find('전 ') ; tmp_str_1 = tmp_str_0 [T1+2 : ] ; 
        T2 = tmp_str_1.find('조회수 ') ; tmp_str_2 = tmp_str_1[T2 : ] ; run_time = tmp_str_1[:T2];
        T3 = run_time.find('시간'); 
        if(T3 == -1):
            continue;
        hour=run_time[:T3]
        if("1" in hour):
            continue
  
        d = sour['title']
        arr_txt.append(d)    

print(arr_txt)


# In[6]:


wb = openpyxl.Workbook()
wb.save('test.xlsx')

sheet = wb.active

sheet['A1'] = 'id'
sheet['B1'] = 'title'

i=1
for a in arr_txt:
    sheet.cell(row=i+1, column=1).value = i
    sheet.cell(row=i+1, column=2).value = a
    i += 1
wb.save('test.xlsx')

